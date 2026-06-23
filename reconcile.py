"""Excel Reconciliation Tool.

Compare two Excel files (e.g. CRM orders vs. bank payments) by a shared key
column and produce a single, formatted Excel report that highlights every
discrepancy, summarises the results, and embeds a bar chart.

Each row from the two inputs is categorised as one of:

* ``MATCHED``            - key present in both files and amounts agree
* ``AMOUNT_MISMATCH``    - key present in both files but amounts differ
* ``MISSING_IN_TARGET``  - key present in the source but not in the target
* ``EXTRA_IN_TARGET``    - key present in the target but not in the source

Run ``python reconcile.py --help`` for usage.
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

# --------------------------------------------------------------------------- #
# Categories and presentation constants
# --------------------------------------------------------------------------- #

# Category labels are also used as dict keys throughout, so they live in one
# place to avoid typos drifting between the reconcile and reporting stages.
MATCHED = "MATCHED"
AMOUNT_MISMATCH = "AMOUNT_MISMATCH"
MISSING_IN_TARGET = "MISSING_IN_TARGET"
EXTRA_IN_TARGET = "EXTRA_IN_TARGET"

# One colour per category, reused for both the worksheet fills and the chart
# bars so the report reads as a single colour key.
COLORS: Dict[str, str] = {
    MATCHED: "C6EFCE",            # light green
    AMOUNT_MISMATCH: "FFEB9C",    # light yellow
    MISSING_IN_TARGET: "FFC7CE",  # light red
    EXTRA_IN_TARGET: "FFD9A0",    # light orange
}

# Solid fills used to colour-code each result sheet. ``PatternFill`` needs the
# colour twice (start/end) for a solid fill in openpyxl.
FILLS: Dict[str, PatternFill] = {
    category: PatternFill("solid", fgColor=color)
    for category, color in COLORS.items()
}

HEADER_FILL = PatternFill("solid", fgColor="305496")  # dark blue
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)
THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)


@dataclass
class ReconResult:
    """Container for the four categorised frames plus headline figures."""

    matched: pd.DataFrame
    mismatches: pd.DataFrame
    missing: pd.DataFrame
    extra: pd.DataFrame
    source_total: float
    target_total: float
    generated_at: datetime

    @property
    def net_difference(self) -> float:
        """Signed gap between the two files (target minus source)."""
        return self.target_total - self.source_total

    @property
    def counts(self) -> Dict[str, int]:
        """Row count per category, in a fixed, display-friendly order."""
        return {
            MATCHED: len(self.matched),
            AMOUNT_MISMATCH: len(self.mismatches),
            MISSING_IN_TARGET: len(self.missing),
            EXTRA_IN_TARGET: len(self.extra),
        }


# --------------------------------------------------------------------------- #
# Core reconciliation logic
# --------------------------------------------------------------------------- #

def load_sheet(path: Path) -> pd.DataFrame:
    """Read the first worksheet of an Excel file into a DataFrame.

    Raises a clear ``SystemExit`` (rather than a raw traceback) when the file
    is missing, so the CLI fails gracefully for non-technical users.
    """
    if not path.exists():
        sys.exit(f"Error: file not found -> {path}")
    return pd.read_excel(path)


def _require_columns(df: pd.DataFrame, columns: list[str], label: str) -> None:
    """Abort early with a readable message if expected columns are absent."""
    missing = [c for c in columns if c not in df.columns]
    if missing:
        sys.exit(
            f"Error: {label} is missing column(s) {missing}. "
            f"Available columns: {list(df.columns)}"
        )


def _one_side(merged: pd.DataFrame, key: str, side: str) -> pd.DataFrame:
    """Slice a suffixed outer-merge back down to a single file's columns.

    After an outer merge with ``suffixes=('_source', '_target')`` every
    overlapping column is duplicated. This rebuilds a clean frame for one side
    and strips the suffix so the output reads like the original input.
    """
    suffix = f"_{side}"
    cols = [key] + [c for c in merged.columns if c.endswith(suffix)]
    out = merged[cols].copy()
    out.columns = [key] + [c[: -len(suffix)] for c in cols[1:]]
    return out.reset_index(drop=True)


def reconcile(
    source: pd.DataFrame,
    target: pd.DataFrame,
    key: str,
    amount_col: str,
    tolerance: float,
) -> ReconResult:
    """Compare two frames by ``key`` and categorise every row.

    Args:
        source: Reference data (e.g. CRM orders) - the source of truth.
        target: Data being checked against the source (e.g. bank payments).
        key: Column used to match rows across the two frames.
        amount_col: Numeric column whose values are compared for matched keys.
        tolerance: Absolute difference below which two amounts count as equal.
            A small tolerance (e.g. 0.01) absorbs floating-point/rounding noise.

    Returns:
        A :class:`ReconResult` holding the four categorised frames and totals.
    """
    _require_columns(source, [key, amount_col], "source")
    _require_columns(target, [key, amount_col], "target")

    # Duplicate keys would create a cartesian blow-up in the merge and make the
    # categories meaningless, so warn loudly rather than silently mis-report.
    for name, df in (("source", source), ("target", target)):
        dupes = df[key].duplicated().sum()
        if dupes:
            print(f"Warning: {dupes} duplicate key(s) found in {name}.")

    # A single outer merge with an indicator column gives us all four
    # categories in one pass; matched keys are then split by amount equality.
    merged = source.merge(
        target,
        on=key,
        how="outer",
        suffixes=("_source", "_target"),
        indicator=True,
    )

    both = merged[merged["_merge"] == "both"].copy()
    diff = both[f"{amount_col}_target"] - both[f"{amount_col}_source"]
    is_match = diff.abs() <= tolerance

    # MATCHED -> keep the source view; the amounts agree by definition.
    matched = _one_side(both[is_match], key, "source")

    # AMOUNT_MISMATCH -> show both amounts side by side plus the delta so a
    # reviewer can act without cross-referencing the source files.
    mism = both[~is_match].copy()
    mismatches = pd.DataFrame({key: mism[key].to_numpy()})
    for col in source.columns:
        if col in (key, amount_col):
            continue
        mismatches[col] = mism[f"{col}_source"].to_numpy()
    mismatches[f"{amount_col}_source"] = mism[f"{amount_col}_source"].to_numpy()
    mismatches[f"{amount_col}_target"] = mism[f"{amount_col}_target"].to_numpy()
    mismatches["difference"] = (
        mism[f"{amount_col}_target"] - mism[f"{amount_col}_source"]
    ).to_numpy()

    missing = _one_side(merged[merged["_merge"] == "left_only"], key, "source")
    extra = _one_side(merged[merged["_merge"] == "right_only"], key, "target")

    return ReconResult(
        matched=matched,
        mismatches=mismatches,
        missing=missing,
        extra=extra,
        source_total=float(source[amount_col].sum()),
        target_total=float(target[amount_col].sum()),
        generated_at=datetime.now(),
    )


# --------------------------------------------------------------------------- #
# Report writing (openpyxl)
# --------------------------------------------------------------------------- #

def _autosize(ws: Worksheet) -> None:
    """Approximate Excel's 'autofit' by widening columns to their content."""
    for column_cells in ws.columns:
        length = max((len(str(c.value)) for c in column_cells if c.value), default=0)
        letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[letter].width = min(length + 4, 40)


def _write_data_sheet(
    wb: Workbook, title: str, df: pd.DataFrame, fill: PatternFill
) -> None:
    """Write one categorised frame to its own colour-coded sheet."""
    ws = wb.create_sheet(title)

    if df.empty:
        ws["A1"] = f"No rows in this category ({title})."
        ws["A1"].font = Font(italic=True, color="808080")
        return

    # ``dataframe_to_rows`` streams header + body; index=False keeps it tidy.
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = THIN_BORDER
            if r_idx == 1:  # header
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center")
            else:  # data rows carry the category's highlight colour
                cell.fill = fill

    ws.freeze_panes = "A2"  # keep the header visible while scrolling
    _autosize(ws)


def _write_summary_sheet(wb: Workbook, result: ReconResult) -> None:
    """Build the Summary sheet: headline figures, counts table, and a chart."""
    ws = wb.create_sheet("Summary", 0)  # first tab

    ws["A1"] = "Reconciliation Summary"
    ws["A1"].font = TITLE_FONT
    ws["A3"] = "Generated:"
    ws["B3"] = result.generated_at.strftime("%Y-%m-%d %H:%M:%S")
    ws["A4"] = "Source total:"
    ws["B4"] = round(result.source_total, 2)
    ws["A5"] = "Target total:"
    ws["B5"] = round(result.target_total, 2)
    ws["A6"] = "Net difference (target - source):"
    ws["B6"] = round(result.net_difference, 2)
    for row in range(3, 7):
        ws[f"A{row}"].font = Font(bold=True)

    # Counts table - also the data source for the embedded chart.
    header_row = 8
    ws.cell(header_row, 1, "Category").font = Font(bold=True)
    ws.cell(header_row, 2, "Count").font = Font(bold=True)
    for offset, (category, count) in enumerate(result.counts.items(), start=1):
        r = header_row + offset
        label = ws.cell(r, 1, category)
        ws.cell(r, 2, count)
        label.fill = FILLS[category]  # mirror each sheet's colour key

    # Bar chart of counts per category, anchored to the right of the table.
    chart = BarChart()
    chart.title = "Rows per category"
    chart.type = "col"
    chart.legend = None
    # No y-axis title: in this small chart the rotated label overlaps the tick
    # numbers, and the on-bar data labels already make the counts obvious.
    # Force both axes to render; openpyxl otherwise sometimes drops the category
    # (x) axis, leaving unlabelled bars.
    chart.x_axis.delete = False
    chart.y_axis.delete = False

    data = Reference(ws, min_col=2, min_row=header_row, max_row=header_row + 4)
    cats = Reference(ws, min_col=1, min_row=header_row + 1, max_row=header_row + 4)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # Print only the count on top of each bar; the category is already on the
    # x-axis, so suppress the series/category names to avoid cluttered labels.
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.showSerName = False
    chart.dataLabels.showCatName = False
    chart.dataLabels.showLegendKey = False

    # Colour each bar to match its category's highlight in the table above.
    series = chart.series[0]
    for idx, category in enumerate(result.counts):
        point = DataPoint(idx=idx)
        point.graphicalProperties.solidFill = COLORS[category]
        series.data_points.append(point)

    chart.height = 7
    chart.width = 14
    ws.add_chart(chart, "D8")

    _autosize(ws)


def write_report(result: ReconResult, output: Path) -> None:
    """Assemble the multi-sheet workbook and save it to ``output``."""
    wb = Workbook()
    # Workbook starts with one blank default sheet; drop it so our named sheets
    # are the only tabs.
    wb.remove(wb.active)

    _write_summary_sheet(wb, result)
    _write_data_sheet(wb, "Matched", result.matched, FILLS[MATCHED])
    _write_data_sheet(wb, "Mismatches", result.mismatches, FILLS[AMOUNT_MISMATCH])
    _write_data_sheet(wb, "Missing", result.missing, FILLS[MISSING_IN_TARGET])
    _write_data_sheet(wb, "Extra", result.extra, FILLS[EXTRA_IN_TARGET])

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #

def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    """Define and parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description="Reconcile two Excel files by a key column and emit a "
        "highlighted, charted Excel report.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--source", required=True, type=Path,
                        help="Source/reference Excel file (e.g. CRM orders).")
    parser.add_argument("--target", required=True, type=Path,
                        help="Target Excel file to check (e.g. bank payments).")
    parser.add_argument("--key", required=True,
                        help="Column name used to match rows across files.")
    parser.add_argument("--output", required=True, type=Path,
                        help="Path for the generated .xlsx report.")
    # The amount column and tolerance are configurable so the tool works for
    # any dataset, not just the bundled sample.
    parser.add_argument("--amount-col", default="amount",
                        help="Numeric column compared for matched keys.")
    parser.add_argument("--tolerance", type=float, default=0.01,
                        help="Max absolute amount gap still treated as a match.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> None:
    """Entry point: load, reconcile, write, and print a short summary."""
    args = parse_args(argv)

    source = load_sheet(args.source)
    target = load_sheet(args.target)

    result = reconcile(
        source=source,
        target=target,
        key=args.key,
        amount_col=args.amount_col,
        tolerance=args.tolerance,
    )
    write_report(result, args.output)

    # Console recap so the user gets feedback without opening the file.
    print(f"Report written to {args.output}")
    for category, count in result.counts.items():
        print(f"  {category:<18} {count}")
    print(f"  Net difference     {result.net_difference:+.2f}")


if __name__ == "__main__":
    main()
